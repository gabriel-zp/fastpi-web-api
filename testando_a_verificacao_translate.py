import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from tkinter import filedialog, Tk, simpledialog
import os
import sys
import io
from datetime import datetime
from fastapi import UploadFile

async def filtrar_e_distribuir_casos(caminho_df_main, num_analistas):

    contents = await caminho_df_main.read()

    df_main = pd.read_excel(io.BytesIO(contents))

    df_main = df_main.iloc[:, 3:]
    df_main = df_main.iloc[:, :14]

    coluna_enquadramento = df_main.columns.get_loc('Enquadramento')
    condicao = (df_main.iloc[:, coluna_enquadramento] == "1.2 D - Acúmulo de Representantes") | (df_main.iloc[:, coluna_enquadramento] == "2.13 M - Envio Vários Clientes p/ mesma Contraparte") | (df_main.iloc[:, coluna_enquadramento] == '3.6 M - Recebimento de varios Sender')
    casos_filtrados = df_main[condicao]

    timestamp = datetime.now().strftime("%Y%m%d%H")

    caminho_planilha1_consolidada = os.path.join("/temp", f"Planilha1(Consolidada)_{timestamp}.xlsx")
    casos_filtrados.to_excel(caminho_planilha1_consolidada, index=False)


    df = casos_filtrados.copy() 
    df.sort_values(by='dc_alerta', inplace=True)

    analistas = ["Analista 1", "Analista 2", "Analista 3", "Analista 4"]
    mapeamento = {}
    analista_atual = analistas[0]

    for nome in df['dc_alerta'].unique():
        mapeamento[nome] = analista_atual
        analista_atual = analistas[(analistas.index(analista_atual) + 1) % len(analistas)]

    df['Analista'] = df['dc_alerta'].map(mapeamento)

    wb = Workbook()

    for analista in range(num_analistas):
        sub_df = df[df['Analista'] == f"Analista {analista+1}"]
        sub_df.drop(columns=['Analista'], inplace=True)
        sub_df.reset_index(drop=True, inplace=True)
    
        sheet = wb.create_sheet(title=f"Analista {analista+1}")
    
        for r_idx, row in enumerate(dataframe_to_rows(sub_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)

    del wb['Sheet']

    segunda_condicao = (df_main.iloc[:, 11] == '2.16 M - Transações de Correspondentes') | (df_main.iloc[:, 11] == '2.15 M - Transações de Funcionários') | (df_main.iloc[:, 11] == '2.05 - Transações de clientes monitorados') 
    segundo_casos_filtrados = df_main[segunda_condicao] # Planilha2(Consolidada)

    caminho_planilha2_consolidada = os.path.join("/temp", f"Planilha2(consolidada)_{timestamp}.xlsx")
    segundo_casos_filtrados.to_excel(caminho_planilha2_consolidada, index=False)

    sheet_n2 = wb.create_sheet(title='N2')

    for caso in dataframe_to_rows(segundo_casos_filtrados, index=False, header=True):
        sheet_n2.append(caso)

    caminho_planilha1_separada = os.path.join("/tmep", f"Planilha1(Separada)_{timestamp}.xlsx")
    wb.save(caminho_planilha1_separada)

    # return wb

# def segunda_verificacao(caminho):

    
    dados_CPFCNPJ = []

    df_filtrado = df['Analista']
    df_filtrado.drop(columns=['Analista'], inplace=True)
    df_filtrado = df_filtrado.iloc[:, 3:]
    df_filtrado = df_filtrado.iloc[:, :14]

    df_principal = df_main.copy()  

    with pd.ExcelFile(wb) as xls:
        for sheet_name in xls.sheet_names:
            df_filtrado = pd.read_excel(xls, sheet_name=sheet_name)

            for index, row in df_filtrado.iterrows():
                dados_CPFCNPJ.append(row['CPF/CNPJ'])

    indices_para_remover = []

    for index, row in df_principal.iterrows():
        if(row['CPF/CNPJ'] in dados_CPFCNPJ):
                indices_para_remover.append(index)

    df_principal.drop(indices_para_remover, inplace=True)
    # df_principal = df_principal.iloc[:, 3:]
    # df_principal = df_principal.iloc[:, :14] 

    df_filtrado_tir = df_principal[df_principal['Produto'] == 'TIR'].drop_duplicates(subset='Cód. Sircoi')

    df_filtrado_n_tir = df_principal[df_principal['Produto'] == 'Câmbio'].drop_duplicates(subset='CPF/CNPJ')

    df_principal = pd.concat([df_filtrado_tir, df_filtrado_n_tir])

    print(df_principal)

    # caminho_planilha2_consolidada = os.path.join(caminho_para_salvar, "Planilha2(consolidada).xlsx")
    df_principal.to_excel(f"temp/Planilha2(consolidada)_{timestamp}.xlsx", index=False)


root = Tk()
root.withdraw()

numero_analistas_input = simpledialog.askstring("input", "Número de analistas a serem separados: ")

caminho_df_main = filedialog.askopenfilename(title="Selecione o arquivo df_main", filetypes=[("Excel Files", "*.xls;*.xlsx")])
caminho_para_salvar = filedialog.askdirectory()

if not caminho_df_main:
    print("Nenhum arquivo selecionado. Encerrando.")
    exit()

caminho_planilha_final = filtrar_e_distribuir_casos(caminho_df_main, int(numero_analistas_input))
# segunda_verificacao(caminho_planilha_final)

df_main_directory = os.path.dirname(caminho_df_main)
caminho_arquivo_final = os.path.join(caminho_para_salvar, "Planilha1(separada).xlsx")

os.rename(caminho_planilha_final, caminho_arquivo_final)
print(f"Processamento concluído. Resultado salvo em: {caminho_arquivo_final}")
