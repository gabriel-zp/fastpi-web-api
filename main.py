from fastapi import FastAPI, UploadFile, File, Form
from fastapi.staticfiles import StaticFiles
import pandas as pd
import io
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook


app = FastAPI()

app.mount("/temp", StaticFiles(directory="temp"), name="temp")

@app.get("/api/v1/endpoint1")
async def endpoint1():
    # Implemente a lógica do seu endpoint aqui
    return {"message": "Você acessou via get o endpoint1"}

@app.post("/api/v1/endpoint1")
async def endpoint1_post(positiveInteger: int = Form(...), excelFile: UploadFile = File(...)):

    contents = await excelFile.read()

    df_tir_cambio = pd.read_excel(io.BytesIO(contents))

    df_tir_cambio.sort_values(by='Produto')


    contador_tir = (df_tir_cambio['Produto'] == 'TIR').sum()
    contador_cambio = (df_tir_cambio['Produto'] == 'Câmbio').sum()

    tir_por_analista = round(contador_tir / positiveInteger)
    cambio_por_analista = round(contador_cambio / positiveInteger)

    df_analista = []

    for analista in range(positiveInteger):
        df_analista.append(pd.concat([df_tir_cambio.iloc[tir_por_analista*analista:tir_por_analista*(analista+1)], 
                              df_tir_cambio.iloc[contador_tir+cambio_por_analista*analista:contador_tir+cambio_por_analista*(analista+1)]]))
        print(df_analista[analista])
        print('----------------')
    
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")

    outputFileName = f"/temp/Planilha2(separada)_{timestamp}.xlsx"

    with pd.ExcelWriter(outputFileName, engine='openpyxl') as writer:
        for analista in range(positiveInteger):
            df_analista[analista].to_excel(writer, sheet_name=f"Analista {analista+1}", index=False)
    
    return {"links": outputFileName}


@app.get("/api/v1/endpoint2")
async def endpoint2():  
    return {"message": "Você acessou o endpoint2"}

@app.post("/api/v1/endpoint2")
async def endpoint2(n: int = Form(...), arquivo_bytes: UploadFile = File(...)):

    contents = await arquivo_bytes.read()

    df_main = pd.read_excel(io.BytesIO(contents))

    df_main = df_main.iloc[:, 3:]
    df_main = df_main.iloc[:, :14]

    coluna_enquadramento = df_main.columns.get_loc('Enquadramento')
    condicao = (df_main.iloc[:, coluna_enquadramento] == "1.2 D - Acúmulo de Representantes") | (df_main.iloc[:, coluna_enquadramento] == "2.13 M - Envio Vários Clientes p/ mesma Contraparte") | (df_main.iloc[:, coluna_enquadramento] == '3.6 M - Recebimento de varios Sender')
    casos_filtrados = df_main[condicao]

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")

    casos_filtrados.to_excel(f"temp/Planilha1(Consolidada)_{timestamp}.xlsx", index=False)


    df = casos_filtrados.copy() 
    df.sort_values(by='dc_alerta', inplace=True)

    analistas = ["Analista 1", "Analista 2", "Analista 3", "Analista 4"]
    mapeamento = {}
    analista_atual = analistas[0]

    for nome in df['dc_alerta'].unique():
        mapeamento[nome] = analista_atual
        analista_atual = analistas[(analistas.index(analista_atual) + 1) % len(analistas)]

    df['Analista'] = df['dc_alerta'].map(mapeamento)

    wb = Workbook()

    for analista in range(n):
        sub_df = df[df['Analista'] == f"Analista {analista+1}"]
        sub_df.drop(columns=['Analista'], inplace=True)
        sub_df.reset_index(drop=True, inplace=True)
    
        sheet = wb.create_sheet(title=f"Analista {analista+1}")
    
        for r_idx, row in enumerate(dataframe_to_rows(sub_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)

    del wb['Sheet']

    segunda_condicao = (df_main.iloc[:, 8] == '2.16 M - Transações de Correspondentes') | (df_main.iloc[:, 8] == '2.15 M - Transações de Funcionários') | (df_main.iloc[:, 8] == '2.05 - Transações de clientes monitorados') 
    segundo_casos_filtrados = df_main[segunda_condicao] # Planilha2(Consolidada)

    sheet_n2 = wb.create_sheet(title='N2')

    for caso in dataframe_to_rows(segundo_casos_filtrados, index=False, header=True):
        sheet_n2.append(caso)

    wb.save(f"temp/Planilha1(Separada)_{timestamp}.xlsx")

    dados_CPFCNPJ = []

    df_filtrado = df_main[(df_main.iloc[:, coluna_enquadramento] == "1.2 D - Acúmulo de Representantes") | (df_main.iloc[:, coluna_enquadramento] == "2.13 M - Envio Vários Clientes p/ mesma Contraparte") | (df_main.iloc[:, coluna_enquadramento] == '3.6 M - Recebimento de varios Sender') | (df_main.iloc[:, coluna_enquadramento] == '2.16 M - Transações de Correspondentes') | (df_main.iloc[:, coluna_enquadramento] == '2.15 M - Transações de Funcionários') | (df_main.iloc[:, coluna_enquadramento] == '2.05 - Transações de clientes monitorados') ]
    df_filtrado = df_filtrado.iloc[:, 3:]
    df_filtrado = df_filtrado.iloc[:, :14]

    df_principal = df_main.copy()  

    for index, row in df_filtrado.iterrows():
                dados_CPFCNPJ.append(row['CPF/CNPJ'])

    indices_para_remover = []

    for index, row in df_principal.iterrows():
        if(row['CPF/CNPJ'] in dados_CPFCNPJ):
                indices_para_remover.append(index)

    df_principal.drop(indices_para_remover, inplace=True)

    df_filtrado_tir = df_principal[df_principal['Produto'] == 'TIR'].drop_duplicates(subset='Cód. Sircoi')

    df_filtrado_n_tir = df_principal[df_principal['Produto'] == 'Câmbio'].drop_duplicates(subset='CPF/CNPJ')

    df_principal = pd.concat([df_filtrado_tir, df_filtrado_n_tir])

    print(df_principal)

    df_principal.to_excel(f"temp/Planilha2(consolidada)_{timestamp}.xlsx", index=False)

    df_tir_cambio = df_principal.copy()

    df_tir_cambio.sort_values(by='Produto')


    contador_tir = (df_tir_cambio['Produto'] == 'TIR').sum()
    contador_cambio = (df_tir_cambio['Produto'] == 'Câmbio').sum()

    tir_por_analista = round(contador_tir / n)
    cambio_por_analista = round(contador_cambio / n)

    df_analista = []

    for analista in range(n):
        df_analista.append(pd.concat([df_tir_cambio.iloc[tir_por_analista*analista:tir_por_analista*(analista+1)], 
                              df_tir_cambio.iloc[contador_tir+cambio_por_analista*analista:contador_tir+cambio_por_analista*(analista+1)]]))
        print(df_analista[analista])
        print('----------------')

    with pd.ExcelWriter(f"temp/Planilha2(separada)_{timestamp}.xlsx", engine='openpyxl') as writer:
        for analista in range(n):
            df_analista[analista].to_excel(writer, sheet_name=f"Analista {analista+1}", index=False)
    
    return {"planilha1_consolidada": f"temp/Planilha1(Consolidada)_{timestamp}.xlsx",
            "planilha1_separada": f"temp/Planilha1(Separada)_{timestamp}.xlsx",
            "planilha2_consolidada": f"temp/Planilha2(consolidada)_{timestamp}.xlsx",
            "planilha2_separada": f"temp/Planilha2(separada)_{timestamp}.xlsx"
            }

@app.get("/api/v1/endpoint3")
async def endpoint3():
    return {"message":"acessando via get enpoint3"}

# Montar a pasta de arquivos estáticos
app.mount("/", StaticFiles(directory="static", html=True), name="static")
